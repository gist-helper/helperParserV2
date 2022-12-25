import datetime
import json

import openpyxl

from meal import MealWrapper, kind_of_meals, slot_endpoints, slot_filenames_postfix
from util import ComplexEncoder, sanitize_menu


def parse_slot(sheet, eng: int, column_day: int, row_slot: int):
    day = sheet.cell(row=2, column=column_day).value
    menus = ""
    for row_menus in range(slot_endpoints[row_slot][0], slot_endpoints[row_slot][1]):
        menus += (
            sanitize_menu((sheet.cell(row=row_menus, column=column_day).value)) + "\n"
        )
        if row_menus == 20:
            if eng:
                menus += "\n\\Corner\\\n"
            else:
                menus += "\n\\코너\\\n"

    meal_wrapper = MealWrapper()
    meal_wrapper.meal.title = (
        "제2학생회관1층" if not eng else "Student Union Bldg.2 1st floor"
    )
    try:
        meal_date = day.strftime("%Y-%m-%d")
    except AttributeError:
        meal_date = datetime(2022, int(day[0:2]), int(day[4:6]), 00, 00, 00)
    meal_wrapper.meal.meal_date = meal_date
    meal_wrapper.meal.kind_of_meal = kind_of_meals[eng][row_slot]
    meal_wrapper.meal.menu = menus.rstrip("\n")
    default_name = day.strftime("%m_%d") + slot_filenames_postfix[eng][row_slot]
    jsonFile = open(f"./{default_name}", "w+", encoding="utf-8")
    json.dump(
        meal_wrapper.__dict__,
        jsonFile,
        indent=4,
        ensure_ascii=False,
        cls=ComplexEncoder,
    )


if __name__ == "__main__":
    file_path = "./2학생회관.xlsx"

    workbook = openpyxl.load_workbook(file_path)

    for lang in range(2):
        for column_day in range(4, 11):
            for row_slot in range(3):
                parse_slot(workbook.worksheets[lang], lang, column_day, row_slot)
